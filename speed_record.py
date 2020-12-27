import xlsxwriter
import openpyxl

from datetime import datetime
import threading
import smtplib
import os
import time
import csv
import schedule
import google_spreadsheet_api

csv_file_name = "ouput.csv"
excel_file_name = "output.xlsx"
speedtest_file_name = "speedtest_out"

google_sheet_id = '1FjHcBYx0Sdxw1oYggY7nkwdqwjxH5-TPe6pxmmFJ_ns'
google_sheet_name = 'spe.result'


def parse_speedtest_meta(start_time, finish_time, file_path):
    fp = open(file_path, 'r')  
    lines = fp.readlines()

    provider = (lines[1].split(" from ")[1]).split("...")[0]
    server = (lines[4].split(" by ")[1]).split(": ")[0]
    ping = ((lines[4].split(" by ")[1]).split(": ")[1]).split(" ")[0]
    dl_speed = lines[6].split(" ")[1]
    ul_speed = lines[8].split(" ")[1]

    meta = []
    meta.append(start_time)
    meta.append(finish_time)
    meta.append(ping)
    meta.append(dl_speed)
    meta.append(ul_speed)
    meta.append(provider)
    meta.append(server)
    #print(meta)
    return meta

def run_speedtest():
    os.system("/home/pi/.local/bin/speedtest-cli > " + speedtest_file_name)
   
def write_data_to_csv():
    with open(csv_file_name, 'a', newline='') as csv_file:
        fwrite = csv.writer(csv_file)
        fwrite.writerow(['a', 'b', 'c'])
        fwrite.writerow(['a1', 'b1', 'c1'])
        fwrite.writerow(['2a', 'b', 'c'])

def write_data_to_excel(meta):

    column_idx = 1
    row_idx = 1
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")

    if (os.path.exists(excel_file_name) == True):
        wb = openpyxl.load_workbook(excel_file_name)
        if date_str in wb.sheetnames:
            write_sheet = wb[date_str]
            row_idx = write_sheet.max_row + 1
        else:
            wb.create_sheet(date_str)
            write_sheet = wb[date_str]
    else:
        wb = openpyxl.Workbook()
        wb.create_sheet(date_str)
        write_sheet = wb[date_str]
 
    #write_sheet['A' + str(new_row)] = now.strftime("%Y_%m_%d_%H_%M_%S")

    for data in meta:
        write_sheet.cell(row=row_idx, column=column_idx).value = data
        column_idx += 1

    #write_sheet.cell(row=10, column=1).value = "ryan_test"

    wb.save(excel_file_name)

def write_data_to_google_sheet(g_api, meta):
    if not g_api.is_init_done:
        g_api.init(google_sheet_id, google_sheet_name)
    g_api.write_data(meta)


def __task():
    start_time = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
    print("speed-test beginning ... %s" % start_time)
    run_speedtest()
    finish_time = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
    print("speed-test done ... %s" % finish_time)
    meta = parse_speedtest_meta(start_time, finish_time, speedtest_file_name)
    write_data_to_excel(meta)

    sheet_api = google_spreadsheet_api.Sheets_Logging()
    write_data_to_google_sheet(sheet_api, meta)

def main():
    print("__ start ++")

    #schedule.every(2).minutes.do(__task)
    #schedule.every(5).seconds.do(__task)
    #schedule.every().day.at('09:30').do(job1)
    schedule.every().hour.at(":05").do(__task)
    schedule.every().hour.at(":20").do(__task)
    schedule.every().hour.at(":35").do(__task)
    schedule.every().hour.at(":50").do(__task)

    while True:
        schedule.run_pending()
        time.sleep(1)


if __name__ == "__main__":
    main()
    
